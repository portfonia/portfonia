"""Unit tests for `app/services/portfolio_export.py` (issue #331).

Pure serialization tests against a hand-built `PortfolioSnapshot` — no DB,
no HTTP. Router-level tests (endpoint wiring, locale precedence, content
type/disposition) live in `test_portfolio_router.py`.
"""

from __future__ import annotations

import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.portfolio_calculator import HoldingValue, PortfolioSnapshot
from app.services.portfolio_export import (
    EXPORT_COLUMNS,
    portfolio_export_filename,
    render_portfolio_export_md,
    render_portfolio_export_xlsx,
)

_HOLDING = HoldingValue(
    holding_id=uuid.UUID("00000000-0000-0000-0000-0000000000aa"),
    name="Apple",
    ticker="AAPL",
    fund_code=None,
    currency="USD",
    asset_type="stock",
    asset_class="us_equity",
    sector="Technology",
    market="US",
    market_value=Decimal("3000.00"),
    market_value_base=Decimal("3000.00"),
    price_as_of=None,
    pricing_mode="auto",
    broker="IBKR",
    account="Brokerage",
    portfolio="Growth",
    avg_cost=Decimal("250.5"),
    shares=Decimal("10"),
    notes="taxable",
    capture_supported=True,
    cost_basis_base=Decimal("2505.00"),
    unrealized_pnl_base=Decimal("495.00"),
    # Real backend scale (PR #335 review 5103601953): compute_portfolio()'s
    # _ratio() stores a 0..1 fraction, not a percent — 495/2505 = 0.1976,
    # matching pnl_base/cost_basis_base above. A literal "19.76" here would
    # mask the exact bug the review caught: the exporter must scale this one
    # column by 100 before rendering, everything else passes through as-is.
    unrealized_pnl_pct=Decimal("0.1976"),
)


def _snapshot(
    holdings: list[HoldingValue] | None = None, price_as_of_date: date | None = None
) -> PortfolioSnapshot:
    snap = PortfolioSnapshot(base_currency="USD")
    snap.holdings = holdings if holdings is not None else [_HOLDING]
    snap.price_as_of_date = price_as_of_date
    # by_* aggregates deliberately populated here to prove the exporter never
    # reads them (issue #331: export is per-holding rows only).
    snap.by_market = {"US": Decimal("3000.00")}
    snap.by_currency = {"USD": Decimal("3000.00")}
    return snap


def test_export_columns_match_design_contract() -> None:
    assert EXPORT_COLUMNS == (
        "ticker",
        "fund_code",
        "name",
        "market",
        "broker",
        "account",
        "portfolio",
        "asset_class",
        "currency",
        "shares",
        "avg_cost",
        "market_value",
        "market_value_base",
        "cost_basis_base",
        "unrealized_pnl_base",
        "unrealized_pnl_pct",
        "pricing_mode",
        "capture_supported",
    )
    assert "sector" not in EXPORT_COLUMNS
    assert "notes" not in EXPORT_COLUMNS
    assert "holding_id" not in EXPORT_COLUMNS


def test_md_header_carries_as_of_and_base_currency_not_as_data_column() -> None:
    body = render_portfolio_export_md(_snapshot(price_as_of_date=date(2026, 1, 5)), "en")
    lines = body.splitlines()
    assert "2026-01-05" in lines[0]
    assert "USD" in lines[1]
    table_lines = [ln for ln in lines if ln.startswith("|")]
    header_cells = [c.strip() for c in table_lines[0].strip("|").split("|")]
    assert len(header_cells) == len(EXPORT_COLUMNS)
    assert "as_of" not in " ".join(header_cells).lower()
    assert "base_currency" not in " ".join(header_cells).lower()


def test_md_no_price_as_of_date_renders_empty_not_none_string() -> None:
    body = render_portfolio_export_md(_snapshot(price_as_of_date=None), "en")
    first_line = body.splitlines()[0]
    assert "None" not in first_line


def test_md_excludes_by_star_aggregates() -> None:
    body = render_portfolio_export_md(_snapshot(), "en")
    assert "by_market" not in body
    assert "by_currency" not in body


def test_md_row_values() -> None:
    body = render_portfolio_export_md(_snapshot(), "en")
    data_lines = [ln for ln in body.splitlines() if ln.startswith("|")][2:]
    assert len(data_lines) == 1
    cells = [c.strip() for c in data_lines[0].strip("|").split("|")]
    row = dict(zip(EXPORT_COLUMNS, cells, strict=True))
    assert row["ticker"] == "AAPL"
    assert row["fund_code"] == ""
    assert row["name"] == "Apple"
    assert row["broker"] == "IBKR"
    assert row["account"] == "Brokerage"
    assert row["portfolio"] == "Growth"
    assert row["shares"] == "10"
    assert row["avg_cost"] == "250.5"
    assert row["market_value_base"] == "3000"
    assert row["pricing_mode"] == "auto"
    assert row["capture_supported"] == "true"
    assert row["unrealized_pnl_pct"] == "19.76"


def test_unrealized_pnl_pct_is_scaled_to_percent_not_raw_ratio() -> None:
    """PR #335 review 5103601953: compute_portfolio() stores unrealized_pnl_pct
    as a 0..1 ratio (portfolio_calculator.py's _ratio(), quantized to 4dp) —
    every other consumer (the /portfolio table's formatPercent, the overview
    email's :.1%) multiplies by 100 before display. Serializing the raw
    Decimal under a "%"-labeled column understates the real figure 100x."""
    holding = HoldingValue(**{**_HOLDING.__dict__, "unrealized_pnl_pct": Decimal("0.1976")})
    snap = _snapshot(holdings=[holding])

    md_row = render_portfolio_export_md(snap, "en").splitlines()[5]
    md_cells = [c.strip() for c in md_row.strip("|").split("|")]
    assert dict(zip(EXPORT_COLUMNS, md_cells, strict=True))["unrealized_pnl_pct"] == "19.76"

    xlsx = render_portfolio_export_xlsx(snap, "en")
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = next(i for i, r in enumerate(rows) if r[0] == "Ticker")
    data_row = rows[header_row_idx + 1]
    pct_idx = EXPORT_COLUMNS.index("unrealized_pnl_pct")
    assert data_row[pct_idx] == pytest.approx(19.76)


def test_md_locale_switches_headers() -> None:
    en = render_portfolio_export_md(_snapshot(), "en")
    zh = render_portfolio_export_md(_snapshot(), "zh")
    assert "Ticker" in en
    assert "代码" in zh
    assert "Ticker" not in zh


def test_md_unrecognized_locale_falls_back_to_english() -> None:
    body = render_portfolio_export_md(_snapshot(), "zh-Hant")
    assert "Ticker" in body


def test_md_empty_holdings_has_header_but_no_data_rows() -> None:
    body = render_portfolio_export_md(_snapshot(holdings=[]), "en")
    table_lines = [ln for ln in body.splitlines() if ln.startswith("|")]
    assert len(table_lines) == 2  # header + separator only


def test_md_pipe_in_field_is_escaped_not_a_column_break() -> None:
    import re

    holding = HoldingValue(**{**_HOLDING.__dict__, "name": "A | B"})
    body = render_portfolio_export_md(_snapshot(holdings=[holding]), "en")
    data_lines = [ln for ln in body.splitlines() if ln.startswith("|")][2:]
    assert len(data_lines) == 1
    assert "A \\| B" in data_lines[0]
    # split on unescaped pipes only, mirroring how a GFM table parser
    # (which treats "\|" as a literal, non-separating pipe) would count cells
    cells = re.split(r"(?<!\\)\|", data_lines[0].strip("|"))
    assert len(cells) == len(EXPORT_COLUMNS)


def test_xlsx_is_valid_and_readable() -> None:
    content = render_portfolio_export_xlsx(_snapshot(price_as_of_date=date(2026, 1, 5)), "en")
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    # metadata rows first (as-of, base_currency), then a header row whose
    # cell count matches EXPORT_COLUMNS, then one data row per holding.
    header_row_idx = next(i for i, r in enumerate(rows) if r[0] == "Ticker")
    header_row = rows[header_row_idx]
    assert len([c for c in header_row if c is not None]) == len(EXPORT_COLUMNS)
    data_row = rows[header_row_idx + 1]
    assert data_row[0] == "AAPL"
    meta_text = " ".join(str(c) for r in rows[:header_row_idx] for c in r if c is not None)
    assert "2026-01-05" in meta_text
    assert "USD" in meta_text


def test_xlsx_excludes_by_star_aggregates() -> None:
    content = render_portfolio_export_xlsx(_snapshot(), "en")
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    assert ws is not None
    all_text = " ".join(
        str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None
    )
    assert "by_market" not in all_text
    assert "by_currency" not in all_text


def test_xlsx_decimal_becomes_numeric_not_string() -> None:
    content = render_portfolio_export_xlsx(_snapshot(), "en")
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = next(i for i, r in enumerate(rows) if r[0] == "Ticker")
    data_row = rows[header_row_idx + 1]
    shares_idx = EXPORT_COLUMNS.index("shares")
    assert isinstance(data_row[shares_idx], int | float)


def test_xlsx_locale_switches_headers() -> None:
    en = render_portfolio_export_xlsx(_snapshot(), "en")
    zh = render_portfolio_export_xlsx(_snapshot(), "zh")
    wb_zh = load_workbook(BytesIO(zh))
    ws_zh = wb_zh.active
    assert ws_zh is not None
    all_text_zh = " ".join(
        str(c) for row in ws_zh.iter_rows(values_only=True) for c in row if c is not None
    )
    assert "代码" in all_text_zh

    wb_en = load_workbook(BytesIO(en))
    ws_en = wb_en.active
    assert ws_en is not None
    all_text_en = " ".join(
        str(c) for row in ws_en.iter_rows(values_only=True) for c in row if c is not None
    )
    assert "Ticker" in all_text_en


def test_columns_identical_between_md_and_xlsx() -> None:
    md = render_portfolio_export_md(_snapshot(), "en")
    md_header = [c.strip() for c in md.splitlines()[3].strip("|").split("|")]

    xlsx = render_portfolio_export_xlsx(_snapshot(), "en")
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = next(i for i, r in enumerate(rows) if r[0] == "Ticker")
    xlsx_header = [c for c in rows[header_row_idx] if c is not None]

    assert md_header == xlsx_header


def test_filename_extension_matches_format() -> None:
    from datetime import UTC, datetime

    now = datetime(2026, 1, 5, 12, 30, 45, tzinfo=UTC)
    assert portfolio_export_filename("md", now) == "portfolio-20260105-123045Z.md"
    assert portfolio_export_filename("xlsx", now) == "portfolio-20260105-123045Z.xlsx"
