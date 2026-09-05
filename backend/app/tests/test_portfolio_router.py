"""Integration tests for /portfolio endpoints."""

from __future__ import annotations

import uuid
from datetime import date
from decimal import Decimal
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.models.fx_rate import FxRate
from app.models.holding import Holding
from app.models.user import User
from app.tests.conftest import seed_user

_USER = uuid.UUID("00000000-0000-0000-0000-000000000001")
_FX_DATE = date(2026, 1, 2)


def _seed(db_session: Session) -> None:
    seed_user(db_session, _USER)
    db_session.add_all(
        [
            FxRate(pair="USDCNY", rate=Decimal("7.0"), rate_date=_FX_DATE, source="test"),
            FxRate(pair="USDHKD", rate=Decimal("8.0"), rate_date=_FX_DATE, source="test"),
            Holding(
                user_id=_USER,
                name="Apple",
                pricing_mode="auto",
                ticker="AAPL",
                currency="USD",
                shares=Decimal("10"),
                market_price=Decimal("300"),
                asset_type="stock",
                sector="Technology",
            ),
        ]
    )
    db_session.flush()


def test_summary_returns_distributions(app_client: TestClient, db_session: Session) -> None:
    _seed(db_session)
    resp = app_client.get("/portfolio/summary")
    assert resp.status_code == 200
    body = resp.json()
    assert body["base_currency"] == "USD"
    assert body["fx_rates_as_of"] == {}
    assert body["total_base"] == "3000.00"
    assert body["by_sector"] == {"Technology": "3000.00"}
    assert body["concentration"]["top_holding_name"] == "Apple"
    assert body["concentration"]["single_holding_high"] is True


def test_summary_defaults_to_principals_own_report_currency_when_omitted(
    app_client: TestClient, db_session: Session
) -> None:
    """Issue #350 item 1: an OMITTED base_currency query param resolves to
    the caller's own persisted users.base_currency, not a hardcoded "USD"
    default — this is what seeds the frontend CurrencySwitcher's initial
    value on first page load."""
    _seed(db_session)
    row = db_session.get(User, _USER)
    assert row is not None
    row.base_currency = "CNY"
    db_session.flush()

    resp = app_client.get("/portfolio/summary")

    assert resp.status_code == 200
    assert resp.json()["base_currency"] == "CNY"
    assert resp.json()["total_base"] == "21000.00"


def test_summary_base_currency_cny(app_client: TestClient, db_session: Session) -> None:
    _seed(db_session)
    resp = app_client.get("/portfolio/summary?base_currency=CNY")
    assert resp.status_code == 200
    assert resp.json()["total_base"] == "21000.00"


def test_summary_rejects_invalid_currency(app_client: TestClient, db_session: Session) -> None:
    _seed(db_session)
    resp = app_client.get("/portfolio/summary?base_currency=XXX")
    assert resp.status_code == 422


def test_summary_accepts_every_valid_currency(app_client: TestClient, db_session: Session) -> None:
    """issue #320: base_currency widened from a 3-value Literal to all 15
    VALID_CURRENCIES entries — EUR/GBP/etc. used to 422 here."""
    from app.schemas.holdings import VALID_CURRENCIES

    _seed(db_session)
    for currency in VALID_CURRENCIES:
        resp = app_client.get(f"/portfolio/summary?base_currency={currency}")
        assert resp.status_code == 200, currency


def test_base_currency_literal_matches_valid_currencies_exactly() -> None:
    """Drift guard: the router's Literal is a hand-copied mirror of
    VALID_CURRENCIES (kept as a Literal, not built dynamically, so mypy
    --strict can verify it) — pin the two together."""
    from typing import get_args

    from app.routers.portfolio import BaseCurrency
    from app.schemas.holdings import VALID_CURRENCIES

    assert set(get_args(BaseCurrency)) == VALID_CURRENCIES


def test_summary_includes_group_account_and_pnl_totals(
    app_client: TestClient, db_session: Session
) -> None:
    """issue #320: by_group/by_broker/P&L totals/price_as_of_date pass
    through the router unchanged from compute_portfolio(); issue #330 added
    a distinct by_account (keyed on Holding.account, not Holding.broker)."""
    _seed(db_session)
    resp = app_client.get("/portfolio/summary")
    assert resp.status_code == 200
    body = resp.json()
    assert body["by_group"] == {"Ungrouped": "3000.00"}
    assert body["by_broker"] == {"Other": "3000.00"}
    assert body["by_account"] == {"Other": "3000.00"}
    assert body["total_cost_basis_base"] == "0"
    assert body["total_unrealized_pnl_base"] == "0"
    assert body["total_unrealized_pnl_pct"] is None
    assert body["price_as_of_date"] is None
    hv = body["holdings"][0]
    assert hv["pricing_mode"] == "auto"
    assert hv["capture_supported"] is True
    assert hv["cost_basis_base"] is None


def test_summary_passes_through_notes(app_client: TestClient, db_session: Session) -> None:
    """Grok review round 2 (PR #322): notes was added to HoldingValueOut to
    close a gap against issue #320 decision 3 / comment 2."""
    seed_user(db_session, _USER)
    db_session.add(
        Holding(
            user_id=_USER,
            name="Private Fund",
            pricing_mode="auto",
            currency="GBP",
            asset_type="stock",
            market="Other",
            capture_supported=False,
            notes="No public ticker",
        )
    )
    db_session.flush()

    resp = app_client.get("/portfolio/summary")

    assert resp.status_code == 200
    assert resp.json()["holdings"][0]["notes"] == "No public ticker"


# POST /refresh moved to POST /admin/portfolio/refresh (issue #128 Ring 1
# stage B, checkpoint B2, decision point 8/11) — an ordinary user must not be
# able to trigger a global market-data refresh. See test_admin_router.py for
# the orchestration test and test_old_portfolio_refresh_route_removed for the
# 404 assertion on this old path.


# ---------------------------------------------------------------------------
# POST /portfolio/send-overview (issue #202) — explicit "Send holdings
# overview" button, not a formal report. Actual send is mocked at the
# Celery-dispatch layer by conftest's autouse `_no_external_notifications`
# (send_portfolio_overview_email_task.delay); these tests only cover the
# cooldown gate and response shape.
# ---------------------------------------------------------------------------


def test_send_overview_first_click_sends(app_client: TestClient, db_session: Session) -> None:
    _seed(db_session)
    resp = app_client.post("/portfolio/send-overview")
    assert resp.status_code == 200
    assert resp.json() == {"sent": True, "retry_after_seconds": None}


def test_send_overview_second_click_within_cooldown_does_not_send(
    app_client: TestClient, db_session: Session
) -> None:
    _seed(db_session)
    first = app_client.post("/portfolio/send-overview")
    assert first.json()["sent"] is True

    second = app_client.post("/portfolio/send-overview")
    body = second.json()
    assert body["sent"] is False
    assert body["retry_after_seconds"] is not None
    assert 0 < body["retry_after_seconds"] <= 900


def test_send_overview_respects_base_currency_param(
    app_client: TestClient, db_session: Session
) -> None:
    _seed(db_session)
    resp = app_client.post("/portfolio/send-overview?base_currency=CNY")
    assert resp.status_code == 200
    assert resp.json()["sent"] is True


def test_send_overview_enqueue_failure_releases_cooldown_for_immediate_retry(
    app_client: TestClient, db_session: Session, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Review 5100733033 leftover: a `.delay()` failure must not leave the
    user locked out for the full 15-minute window over a send that never
    actually got queued."""
    _seed(db_session)
    monkeypatch.setattr(
        "app.routers.portfolio.send_portfolio_overview_email_task.delay",
        MagicMock(side_effect=RuntimeError("broker down")),
    )

    first = app_client.post("/portfolio/send-overview")
    assert first.json() == {"sent": False, "retry_after_seconds": None}

    monkeypatch.setattr(
        "app.routers.portfolio.send_portfolio_overview_email_task.delay", MagicMock()
    )
    second = app_client.post("/portfolio/send-overview")
    assert second.json()["sent"] is True


# ---------------------------------------------------------------------------
# GET /portfolio/export (issue #331) — computed-result snapshot export,
# distinct from GET /holdings/export (issue #92/#310, declared/unpriced
# fields for re-import). Column-set/format details are covered by
# test_portfolio_export.py; these tests cover endpoint wiring only: params,
# content-type/disposition, locale precedence, and auth.
# ---------------------------------------------------------------------------


def test_export_md_returns_markdown_file(app_client: TestClient, db_session: Session) -> None:
    _seed(db_session)
    resp = app_client.get("/portfolio/export?format=md")
    assert resp.status_code == 200
    assert "text/markdown" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]
    import re as _re

    assert _re.search(r'filename="portfolio-\d{8}-\d{6}Z\.md"', resp.headers["content-disposition"])
    assert "AAPL" in resp.text


def test_export_xlsx_returns_valid_workbook(app_client: TestClient, db_session: Session) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    _seed(db_session)
    resp = app_client.get("/portfolio/export?format=xlsx")
    assert resp.status_code == 200
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp.headers["content-disposition"]
    import re as _re

    assert _re.search(
        r'filename="portfolio-\d{8}-\d{6}Z\.xlsx"', resp.headers["content-disposition"]
    )
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws is not None
    all_text = " ".join(str(c) for row in ws.iter_rows(values_only=True) for c in row if c)
    assert "AAPL" in all_text


def test_export_rejects_missing_format(app_client: TestClient, db_session: Session) -> None:
    _seed(db_session)
    resp = app_client.get("/portfolio/export")
    assert resp.status_code == 422


def test_export_rejects_invalid_format(app_client: TestClient, db_session: Session) -> None:
    _seed(db_session)
    resp = app_client.get("/portfolio/export?format=csv")
    assert resp.status_code == 422


def test_export_respects_base_currency_param(app_client: TestClient, db_session: Session) -> None:
    _seed(db_session)
    resp = app_client.get("/portfolio/export?format=md&base_currency=CNY")
    assert resp.status_code == 200
    assert "本位币: CNY" in resp.text or "Base currency: CNY" in resp.text


def test_export_uses_report_locale_when_no_locale_param_given(
    app_client: TestClient, db_session: Session
) -> None:
    from app.models.user import User

    _seed(db_session)
    user = db_session.get(User, _USER)
    assert user is not None
    user.locale = "zh"
    db_session.commit()

    resp = app_client.get("/portfolio/export?format=md")
    assert "代码" in resp.text


def test_export_locale_param_takes_precedence_over_report_locale(
    app_client: TestClient, db_session: Session
) -> None:
    from app.models.user import User

    _seed(db_session)
    user = db_session.get(User, _USER)
    assert user is not None
    user.locale = "zh"
    db_session.commit()

    resp = app_client.get("/portfolio/export?format=md&locale=en")
    assert "Ticker" in resp.text
    assert "代码" not in resp.text


def test_export_md_and_xlsx_have_identical_columns(
    app_client: TestClient, db_session: Session
) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    _seed(db_session)
    md = app_client.get("/portfolio/export?format=md&locale=en").text
    md_header = [c.strip() for c in md.splitlines()[3].strip("|").split("|")]

    xlsx = app_client.get("/portfolio/export?format=xlsx&locale=en").content
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = next(i for i, r in enumerate(rows) if r[0] == "Ticker")
    xlsx_header = [c for c in rows[header_row_idx] if c is not None]

    assert md_header == xlsx_header


def test_export_excludes_by_star_aggregates(app_client: TestClient, db_session: Session) -> None:
    _seed(db_session)
    resp = app_client.get("/portfolio/export?format=md")
    assert "by_market" not in resp.text
    assert "by_currency" not in resp.text
